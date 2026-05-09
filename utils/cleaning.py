# -*- coding: utf-8 -*-
"""数据清洗和匹配逻辑"""

import re
from decimal import Decimal, ROUND_HALF_UP
from openpyxl import load_workbook


def _round_half_up(value):
    """四舍五入到整数（传统ROUND_HALF_UP，区别于Python内置banker's rounding）"""
    return int(Decimal(str(value)).quantize(Decimal('1'), rounding=ROUND_HALF_UP))


# ============================================================
# 城市→省份兜底表（使用全称，用于匹配价格表）
# ============================================================
PROVINCE_HINTS = {
    '北京': '北京', '天津': '天津', '上海': '上海', '重庆': '重庆',
    '广州': '广东省', '深圳': '广东省', '东莞': '广东省', '佛山': '广东省',
    '惠州': '广东省', '中山': '广东省', '珠海': '广东省', '江门': '广东省',
    '成都': '四川省', '绵阳': '四川省', '德阳': '四川省', '眉山': '四川省',
    '资阳': '四川省', '乐山': '四川省', '泸州': '四川省', '南充': '四川省',
    '西安': '陕西省', '咸阳': '陕西省', '宝鸡': '陕西省', '延安': '陕西省',
    '武汉': '湖北省', '襄阳': '湖北省', '宜昌': '湖北省', '荆州': '湖北省',
    '长沙': '湖南省', '岳阳': '湖南省', '株洲': '湖南省',
    '南京': '江苏省', '苏州': '江苏省', '无锡': '江苏省', '常州': '江苏省',
    '南通': '江苏省', '徐州': '江苏省', '扬州': '江苏省',
    '杭州': '浙江省', '宁波': '浙江省', '温州': '浙江省', '嘉兴': '浙江省',
    '金华': '浙江省', '台州': '浙江省', '绍兴': '浙江省',
    '郑州': '河南省', '洛阳': '河南省', '新乡': '河南省', '南阳': '河南省',
    '济南': '山东省', '青岛': '山东省', '烟台': '山东省', '潍坊': '山东省',
    '临沂': '山东省', '淄博': '山东省',
    '沈阳': '辽宁省', '大连': '辽宁省', '鞍山': '辽宁省', '锦州': '辽宁省',
    '哈尔滨': '黑龙江省', '大庆': '黑龙江省', '齐齐哈尔': '黑龙江省',
    '长春': '吉林省', '吉林': '吉林省',
    '合肥': '安徽省', '芜湖': '安徽省', '蚌埠': '安徽省',
    '南昌': '江西省', '赣州': '江西省', '九江': '江西省',
    '福州': '福建省', '厦门': '福建省', '泉州': '福建省', '漳州': '福建省',
    '昆明': '云南省', '贵阳': '贵州省', '遵义': '贵州省',
    '南宁': '广西壮族自治区', '桂林': '广西壮族自治区', '柳州': '广西壮族自治区',
    '海口': '海南省', '三亚': '海南省',
    '石家庄': '河北省', '保定': '河北省', '唐山': '河北省', '廊坊': '河北省',
    '太原': '山西省', '大同': '山西省',
    '呼和浩特': '内蒙古自治区', '包头': '内蒙古自治区',
    '兰州': '甘肃省', '天水': '甘肃省',
    '乌鲁木齐': '新疆维吾尔自治区', '克拉玛依': '新疆维吾尔自治区',
    '银川': '宁夏回族自治区', '石嘴山': '宁夏回族自治区',
    '西宁': '青海省', '格尔木': '青海省',
    '拉萨': '西藏自治区',
}

# ============================================================
# 省份名规范化：短名 → 全称（用于匹配价格表 biao_cross）
# ============================================================
PROVINCE_SHORT_TO_FULL = {
    '北京': '北京', '天津': '天津', '上海': '上海', '重庆': '重庆',
    '河北': '河北省', '山西': '山西省', '内蒙古': '内蒙古自治区', '内蒙': '内蒙古自治区',
    '辽宁': '辽宁省', '吉林': '吉林省', '黑龙江': '黑龙江省',
    '江苏': '江苏省', '浙江': '浙江省', '安徽': '安徽省', '福建': '福建省',
    '江西': '江西省', '山东': '山东省', '河南': '河南省', '湖北': '湖北省',
    '湖南': '湖南省', '广东': '广东省', '广西': '广西壮族自治区',
    '海南': '海南省', '四川': '四川省', '贵州': '贵州省', '云南': '云南省',
    '西藏': '西藏自治区', '陕西': '陕西省', '甘肃': '甘肃省', '青海': '青海省',
    '宁夏': '宁夏回族自治区', '新疆': '新疆维吾尔自治区', '台湾': '台湾',
}


def normalize_province(prov):
    """将短省份名转为全称，用于匹配价格表"""
    if not prov:
        return prov
    return PROVINCE_SHORT_TO_FULL.get(str(prov).strip(), str(prov).strip())


# 全称→短名反向映射（自动从 PROVINCE_SHORT_TO_FULL 构建）
PROVINCE_FULL_TO_SHORT = {v: k for k, v in PROVINCE_SHORT_TO_FULL.items()}


# ============================================================
# 到件地区清洗
# ============================================================
def normalize_area(area):
    """
    拆分到件地区。
    输入: '成都/资阳/眉山' 或 '东莞'
    返回: (原始值, 主城市, 城市列表)
    """
    if not area:
        return '', '', []
    text = str(area).strip()
    parts = [p.strip() for p in re.split(r'[/、,，\-]', text) if p.strip()]
    main = parts[0] if parts else text
    return text, main, parts


# ============================================================
# 加载城市→省份映射
# ============================================================
def load_city_province_map(path):
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    mapping = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and row[1]:
            mapping[str(row[0]).strip()] = str(row[1]).strip()
    return mapping


def guess_province(city, city_map):
    if not city:
        return ''
    city = str(city).strip()
    if city in city_map:
        return city_map[city]
    for k, v in PROVINCE_HINTS.items():
        if city.startswith(k):
            return v
    return ''


# ============================================================
# 加载订单明细（快递单号 → 店铺名称）
# ============================================================
def load_order_shop_map(path):
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header = [str(v).strip() if v is not None else '' for v in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    idx = {name: i for i, name in enumerate(header)}

    shop_col    = idx.get('店铺名称', -1)
    express_col = idx.get('快递单号', -1)

    result = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or all(v is None for v in row):
            continue
        if express_col != -1 and express_col < len(row) and row[express_col]:
            key = str(row[express_col]).strip()
            shop = str(row[shop_col]).strip() if (shop_col != -1 and shop_col < len(row) and row[shop_col]) else ''
            result[key] = shop
    return result


# ============================================================
# 加载3张价格表
# ============================================================
def _load_price_sheet(ws):
    """加载单个价格Sheet，返回字典列表"""
    header = [str(v).strip() if v is not None else '' for v in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    idx = {name: i for i, name in enumerate(header)}
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or all(v is None for v in row):
            continue
        item = {col: (row[i] if i < len(row) else None) for col, i in idx.items()}
        rows.append(item)
    return rows


def load_price_table(path):
    """
    加载3个Sheet，返回字典：
    {
        'ganpei': [...],   # 顺丰干配
        'biao_same': [...], # 顺丰标快（同省）
        'biao_cross': [...], # 顺丰标快（大陆地区异地）
    }
    """
    wb = load_workbook(path, data_only=True)
    result = {}
    for sname in wb.sheetnames:
        ws = wb[sname]
        if sname == '顺丰干配':
            result['ganpei'] = _load_price_sheet(ws)
        elif '同省' in sname:
            result['biao_same'] = _load_price_sheet(ws)
        elif '异地' in sname:
            result['biao_cross'] = _load_price_sheet(ws)
    return result


def _city_match_in_list(city, city_list_str):
    """
    判断主城市 city 是否匹配 city_list_str（逗号/斜杠分隔的城市列表）。
    city_list_str 如: '东莞,中山,云浮,佛山...' 或 '成都/资阳/眉山'
    """
    if not city or not city_list_str:
        return False
    city = str(city).strip()
    # 支持 / 和 , 分隔
    parts = [p.strip() for p in re.split(r'[/,]', str(city_list_str)) if p.strip()]
    for part in parts:
        if city == part.strip():
            return True
    return False


def _parse_weight_threshold(threshold_str):
    """
    解析重量阈值字符串，如 '＜30' -> 30, '≥20' -> 20
    返回 None 表示解析失败
    """
    if not threshold_str:
        return None
    s = str(threshold_str).strip()
    if s.startswith('＜'):
        try:
            return float(s[1:])
        except:
            return None
    elif s.startswith('≥'):
        try:
            return float(s[1:])
        except:
            return None
    return None


def calc_freight(price_tables, product, service, sender_area, dest_city, dest_province, weight):
    """
    运费计算主函数。

    参数:
        price_tables: load_price_table() 返回的3个Sheet数据
        product: 产品类型（顺丰干配/顺丰标快/顺丰特快）
        service: 服务类型（运费/转寄退回/保价/包装服务）
        sender_area: 寄件地区（如 "惠州"、"沈阳/铁岭/抚顺"）
        dest_city: 到件主城市（如 "拉萨"、"惠州"）
        dest_province: 到达省份
        weight: 计费重量

    返回 (freight, matched)
        freight: 计算出的运费，匹配不到返回 None
        matched: 是否匹配到价格
    """
    # 保价/包装服务：暂不处理；转寄/退回仅对干配生效（享受0.6折扣）
    if service not in ('运费', ''):
        if not (product == '顺丰干配' and service == '转寄/退回'):
            return None, False

    # 重量解析
    try:
        w = float(weight)
    except (TypeError, ValueError):
        return None, False

    # ========== 顺丰干配 ==========
    if product == '顺丰干配':
        # 内蒙古使用标快异地价格
        if dest_province and normalize_province(dest_province) == '内蒙古自治区':
            biao_cross = price_tables.get('biao_cross', [])
            result = _calc_biao(biao_cross, normalize_province(dest_province), dest_city, w)
            if result[0] is not None:
                return result
            result = _calc_biao_by_city(biao_cross, dest_city, w)
            return result
        return _calc_ganpei(price_tables['ganpei'], dest_province, w, service)

    # ========== 顺丰标快 / 顺丰特快 ==========
    if product in ('顺丰标快', '顺丰特快'):
        # 先获取寄件省份
        sender_main = normalize_area(sender_area)[1] if sender_area else ''
        sender_province_raw = guess_province(sender_main, {})
        if not sender_province_raw:
            return None, False
        sender_province = normalize_province(sender_province_raw)
        dest_province_norm = normalize_province(dest_province)
        # 同省判断（用原始值比较，避免全称/短名不一致问题）
        s_raw = sender_province_raw.strip()
        d_raw = (dest_province or '').strip()
        if (sender_province == dest_province_norm or
                s_raw == d_raw or
                normalize_province(s_raw) == normalize_province(d_raw)):
            # 同省：使用 biao_same（传原始值，由函数内部处理短/全称兼容）
            return _calc_biao(price_tables.get('biao_same', []), sender_province_raw, dest_city, w)
        else:
            # 异地：使用 biao_cross
            # 先用寄件省份查；若寄件省份在biao_cross中不存在，用目的地省份查（兜底1）
            # 若目的地省份也查不到，用目的地城市在biao_cross中搜索匹配行（兜底2）
            biao_cross = price_tables.get('biao_cross', [])
            result = _calc_biao(biao_cross, sender_province, dest_city, w)
            if result[0] is None and dest_province:
                result = _calc_biao(biao_cross, dest_province_norm, dest_city, w)
            if result[0] is None and dest_city:
                result = _calc_biao_by_city(biao_cross, dest_city, w)
            return result

    return None, False


def _calc_ganpei(ganpei_rows, dest_province, weight, service):
    """顺丰干配计费：按到达省份匹配"""
    if not dest_province:
        return None, False
    for row in ganpei_rows:
        if str(row.get('到达省份', '')).strip() != dest_province.strip():
            continue
        try:
            first_w = float(row.get('首重重量（kg）') or 0)
            first_p = float(row.get('首重价格（元）') or 0)
            step_raw = row.get('续重倍率')
            step_p = float(step_raw) if step_raw is not None and str(step_raw).strip() not in ('', '/') else 0
        except (TypeError, ValueError):
            continue

        if weight <= first_w:
            base = first_p
        else:
            base = first_p + (weight - first_w) * step_p

        if str(service).strip() == '转寄/退回':
            freight = base * 0.6
        else:
            freight = base

        return round(freight, 2), True  # 干配保留两位小数

    return None, False


def _calc_biao(biao_rows, sender_province, dest_city, weight):
    """
    顺丰标快/特快计费：
    匹配 寄件省份 + 到件城市，在 city_list 中模糊匹配 dest_city。
    按重量区间分段计算。
    sender_province 可以是短名或全称，函数内部会尝试两种形式匹配。
    """
    if not dest_city or not sender_province:
        return None, False

    s_raw = sender_province.strip()
    # 转换为短名和全称
    if s_raw in PROVINCE_SHORT_TO_FULL:
        s_short, s_full = s_raw, normalize_province(s_raw)
    elif s_raw in PROVINCE_FULL_TO_SHORT:
        s_short, s_full = PROVINCE_FULL_TO_SHORT[s_raw], s_raw
    else:
        s_short, s_full = s_raw, s_raw

    for row in biao_rows:
        row_prov = str(row.get('寄件省份', '')).strip()
        # 兼容：行中可能是短名或全称
        if row_prov != s_short and row_prov != s_full:
            continue
        city_list_str = str(row.get('到件地区/寄件地区', '')).strip()
        if not _city_match_in_list(dest_city, city_list_str):
            continue

        try:
            first_w = float(row.get('首重重量（kg）') or 0)
            first_p = float(row.get('首重价格（元）') or 0)
            step_th1_raw = row.get('续重重量1（kg）')
            step_p1_raw  = row.get('续重单价1（元）')
            step_th2_raw = row.get('续重重量2（kg）')
            step_p2_raw  = row.get('续重单价2（元）')

            step_th1 = _parse_weight_threshold(step_th1_raw) if step_th1_raw else None
            step_p1  = float(step_p1_raw) if step_p1_raw is not None and str(step_p1_raw).strip() not in ('', '/') else 0
            step_th2 = _parse_weight_threshold(step_th2_raw) if step_th2_raw else None
            step_p2  = float(step_p2_raw) if step_p2_raw is not None and str(step_p2_raw).strip() not in ('', '/') else None
        except (TypeError, ValueError):
            continue

        # 计算运费
        if weight <= first_w:
            freight = first_p
        else:
            excess = weight - first_w
            # 判断用哪个续重单价
            step_price = step_p1  # 默认用续重单价1
            if step_th2 is not None and step_p2 is not None:
                # 有第二档重量阈值
                if weight >= step_th2:
                    step_price = step_p2
                elif step_th1 is not None and step_th1 < step_th2:
                    # 在两个阈值之间，用第一档
                    step_price = step_p1
            elif step_th1 is not None:
                # 只有第一档
                if weight >= step_th1:
                    # 如果有第二档但第二档价格为/，继续用第一档
                    if step_p2 is not None:
                        step_price = step_p2

            freight = first_p + excess * step_price

        return _round_half_up(freight), True

    return None, False


def _calc_biao_by_city(biao_rows, dest_city, weight):
    """
    兜底搜索：忽略省份，直接在所有行中查找 dest_city 匹配的行，
    并使用该行的计费规则进行计算。
    """
    if not dest_city:
        return None, False
    for row in biao_rows:
        city_list_str = str(row.get('到件地区/寄件地区', '')).strip()
        if not _city_match_in_list(dest_city, city_list_str):
            continue
        try:
            first_w = float(row.get('首重重量（kg）') or 0)
            first_p = float(row.get('首重价格（元）') or 0)
            step_th1_raw = row.get('续重重量1（kg）')
            step_p1_raw  = row.get('续重单价1（元）')
            step_th2_raw = row.get('续重重量2（kg）')
            step_p2_raw  = row.get('续重单价2（元）')

            step_th1 = _parse_weight_threshold(step_th1_raw) if step_th1_raw else None
            step_p1  = float(step_p1_raw) if step_p1_raw is not None and str(step_p1_raw).strip() not in ('', '/') else 0
            step_th2 = _parse_weight_threshold(step_th2_raw) if step_th2_raw else None
            step_p2  = float(step_p2_raw) if step_p2_raw is not None and str(step_p2_raw).strip() not in ('', '/') else None
        except (TypeError, ValueError):
            continue

        if weight <= first_w:
            freight = first_p
        else:
            excess = weight - first_w
            step_price = step_p1
            if step_th2 is not None and step_p2 is not None:
                if weight >= step_th2:
                    step_price = step_p2
                elif step_th1 is not None and step_th1 < step_th2:
                    step_price = step_p1
            elif step_th1 is not None:
                if weight >= step_th1:
                    if step_p2 is not None:
                        step_price = step_p2
            freight = first_p + excess * step_price

        return _round_half_up(freight), True

    return None, False


# ============================================================
# 加载4张新匹配表：B站 / KOC / 出入库 / 售后
# ============================================================
def load_bill_matching_maps(bizhan_path, koc_path, io_path, aftersale_path):
    return {
        'bizhan':    load_bizhan_map(bizhan_path),
        'koc':       load_koc_map(koc_path),
        'io':        load_io_map(io_path),
        'aftersale': load_aftersale_map(aftersale_path),
    }


def _header_idx(ws, keywords, row_start=1, row_end=5):
    for r in range(row_start, row_end + 1):
        row = [str(v).strip() if v is not None else '' for v in next(ws.iter_rows(min_row=r, max_row=r, values_only=True))]
        if all(any(k in v for v in row) for k in keywords):
            return {name: i for i, name in enumerate(row) if name}
    return None


def load_bizhan_map(path):
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header = None
    header_row = 1
    for i in range(1, 6):
        row = [str(v).strip() if v is not None else '' for v in next(ws.iter_rows(min_row=i, max_row=i, values_only=True))]
        if any('要发的型号' in v or '发出快递单号' in v for v in row):
            header = row
            header_row = i
            break
    if not header:
        return {}

    idx = {name: i for i, name in enumerate(header)}
    sf_col    = idx.get('发出快递单号', -1)
    model_col = idx.get('实际发出型号', -1)

    result = {}
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not row or all(v is None for v in row):
            continue
        if sf_col != -1 and sf_col < len(row) and row[sf_col]:
            key = str(row[sf_col]).strip()
            model = str(row[model_col]).strip() if (model_col != -1 and model_col < len(row) and row[model_col]) else ''
            result[key] = model
    return result


def load_koc_map(path):
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header = None
    header_row = 1
    for i in range(1, 6):
        row = [str(v).strip() if v is not None else '' for v in next(ws.iter_rows(min_row=i, max_row=i, values_only=True))]
        if any('要发的型号' in v or '发出快递单号' in v for v in row):
            header = row
            header_row = i
            break
    if not header:
        return {}

    idx = {name: i for i, name in enumerate(header)}
    send_col    = idx.get('发出快递单号', -1)
    return_col  = idx.get('寄回快递单号', -1)
    model_col   = idx.get('要发的型号', -1)

    result = {}
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not row or all(v is None for v in row):
            continue
        model = str(row[model_col]).strip() if (model_col != -1 and model_col < len(row) and row[model_col]) else ''
        for col in [send_col, return_col]:
            if col != -1 and col < len(row) and row[col]:
                key = str(row[col]).strip()
                if key:
                    result[key] = model
    return result


def load_io_map(path):
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header = [str(v).strip() if v is not None else '' for v in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    idx = {name: i for i, name in enumerate(header)}

    io_col   = idx.get('出入库单号', -1)
    note_col = idx.get('备注', -1)

    result = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or all(v is None for v in row):
            continue
        if io_col != -1 and io_col < len(row) and row[io_col]:
            key = str(row[io_col]).strip()
            note = str(row[note_col]).strip() if (note_col != -1 and note_col < len(row) and row[note_col]) else ''
            result[key] = note
    return result


def load_aftersale_map(path):
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header = None
    header_row = 1
    for i in range(1, 6):
        row = [str(v).strip() if v is not None else '' for v in next(ws.iter_rows(min_row=i, max_row=i, values_only=True))]
        if any('退回快递单号' in v or '退货单号' in v for v in row):
            header = row
            header_row = i
            break
    if not header:
        return {}

    idx = {name: i for i, name in enumerate(header)}
    sf_col   = idx.get('退回快递单号', -1)
    shop_col = idx.get('店铺名称', -1)

    result = {}
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not row or all(v is None for v in row):
            continue
        if sf_col != -1 and sf_col < len(row) and row[sf_col]:
            key = str(row[sf_col]).strip()
            shop = str(row[shop_col]).strip() if (shop_col != -1 and shop_col < len(row) and row[shop_col]) else ''
            result[key] = shop
    return result


# ============================================================
# 级联店铺匹配
# ============================================================
def cascade_shop_match(waybill_no, order_map, match_maps):
    if not waybill_no:
        return '', '', 0

    wb_key = str(waybill_no).strip()

    if wb_key in order_map:
        shop = order_map[wb_key]
        if shop:
            return shop, '', 1

    aftersale = match_maps.get('aftersale', {})
    if wb_key in aftersale:
        shop = aftersale[wb_key]
        if shop:
            return shop, '', 2

    koc = match_maps.get('koc', {})
    if wb_key in koc:
        model = koc[wb_key]
        return '推广KOC', f'要发的型号:{model}', 3

    bizhan = match_maps.get('bizhan', {})
    if wb_key in bizhan:
        model = bizhan[wb_key]
        return 'B站抽奖', f'实际发出型号:{model}', 4

    return '', '', 0


# ============================================================
# 出入库单号&备注匹配
# ============================================================
def match_io_note(shop, waybill_no, order_map, match_maps, existing_remark):
    if not waybill_no:
        return ''

    if shop and '[其他]' in str(shop) and '吉他情报局' in str(shop):
        pass
    else:
        wb_key = str(waybill_no).strip()
        resolved_shop = order_map.get(wb_key, '')
        if not (resolved_shop and '[其他]' in str(resolved_shop) and '吉他情报局' in str(resolved_shop)):
            return ''

    io_map = match_maps.get('io', {})
    io_note = io_map.get(str(waybill_no).strip(), '')
    return f'出入库单号:{waybill_no}; {io_note}' if io_note else f'出入库单号:{waybill_no}'


# ============================================================
# 主处理函数
# ============================================================
def process_bill(bill_path, order_map, city_map, price_tables, match_maps):
    """
    处理账单明细，返回 (workbook, 处理结果列表, 统计摘要)

    新增列：
      店铺 / 省份 / 运费 / 上浮费 / 总运费 / 是否一致？ / 是否异常？ / 备注
    """
    wb = load_workbook(bill_path)
    ws = wb['账单明细']

    header = [cell.value for cell in ws[2]]
    col_idx = {str(v).strip(): i + 1 for i, v in enumerate(header) if v is not None}

    new_cols = ['店铺', '省份', '运费', '上浮费', '总运费', '是否异常？', '是否一致？', '备注']
    start_col = ws.max_column + 1
    for i, col in enumerate(new_cols, start=start_col):
        ws.cell(row=2, column=i, value=col)

    processed = []
    stats = {
        'total': 0, 'abnormal': 0,
        'consistent_ok': 0, 'consistent_fail': 0,
        'matched_shop': 0, 'matched_province': 0, 'matched_price': 0,
        'shop_order': 0, 'shop_aftersale': 0, 'shop_koc': 0, 'shop_bizhan': 0,
    }

    for r in range(3, ws.max_row + 1):
        row_empty = all(ws.cell(r, c).value is None for c in range(1, ws.max_column + 1))
        if row_empty:
            continue

        stats['total'] += 1
        remarks = []

        waybill_no = ws.cell(r, col_idx.get('运单号码', 3)).value
        sender_area = ws.cell(r, col_idx.get('寄件地区', 4)).value  # 新增
        area       = ws.cell(r, col_idx.get('到件地区', 5)).value
        product    = ws.cell(r, col_idx.get('产品类型', 8)).value
        service    = ws.cell(r, col_idx.get('服务', 14)).value
        weight     = ws.cell(r, col_idx.get('计费重量', 7)).value
        fee_orig   = ws.cell(r, col_idx.get('费用(元)', 10)).value   # 对比基准
        total_orig = ws.cell(r, col_idx.get('应付金额', 12)).value  # 输出用

        # 1. 到件地区清洗
        raw_area, main_city, city_list = normalize_area(area)

        # 2. 寄件地区清洗（取主城市）
        sender_main = normalize_area(str(sender_area) if sender_area else '')[1] if sender_area else ''

        # 3. 级联店铺匹配
        shop, extra_note, match_level = cascade_shop_match(waybill_no, order_map, match_maps)
        if match_level == 1:
            stats['shop_order'] += 1
        elif match_level == 2:
            stats['shop_aftersale'] += 1
        elif match_level == 3:
            stats['shop_koc'] += 1
        elif match_level == 4:
            stats['shop_bizhan'] += 1

        # 4. 出入库单号匹配
        io_extra = match_io_note(shop, waybill_no, order_map, match_maps, '')
        if io_extra:
            remarks.append(io_extra)

        # 5. 追加KOC/B站型号备注
        if extra_note:
            remarks.append(extra_note)

        # 6. 省份匹配
        province = guess_province(main_city, city_map)

        # 7. 运费计算（使用新的 calc_freight）
        freight, price_matched = calc_freight(
            price_tables,
            str(product).strip() if product else '',
            str(service).strip() if service else '',
            str(sender_area).strip() if sender_area else '',
            main_city,
            province,
            weight,
        )

        surcharge = None

        shop_ok     = bool(shop)
        province_ok = bool(province)
        price_ok    = freight is not None

        if shop_ok:
            stats['matched_shop'] += 1
        else:
            remarks.append('未匹配店铺')
        if province_ok:
            stats['matched_province'] += 1
        else:
            remarks.append('未匹配省份')
        if price_ok:
            stats['matched_price'] += 1
        else:
            remarks.append('未匹配价格')

        abnormal = '异常' if (not shop_ok or not province_ok or not price_ok) else '正常'
        if abnormal == '异常':
            stats['abnormal'] += 1

        computed_total = freight

        consistent = ''
        if freight is not None and total_orig is not None:
            try:
                freight_f = float(freight)
                total_f = float(str(total_orig).strip())
                if abs(freight_f - total_f) <= 0.01:
                    consistent = '一致'
                    stats['consistent_ok'] += 1
                else:
                    consistent = '不一致'
                    stats['consistent_fail'] += 1
                    remarks.append(f'总运费({freight_f})≠应付({total_f})')
            except (TypeError, ValueError):
                consistent = ''

        if raw_area != main_city and len(city_list) > 1:
            remarks.insert(0, f'到件地区含多城市:{raw_area}')

        remark_text = '; '.join(remarks)
        values = [shop, province, freight, surcharge, computed_total, abnormal, consistent, remark_text]
        for i, val in enumerate(values, start=start_col):
            ws.cell(row=r, column=i, value=val)

        processed.append([
            waybill_no, raw_area, shop, province, freight, surcharge,
            computed_total, abnormal, consistent, remark_text
        ])

    return wb, processed, stats


def export_result(wb, output_path):
    wb.save(output_path)
